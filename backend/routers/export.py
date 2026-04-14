"""엑셀 내보내기 및 CSV 다운로드"""

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from database import get_db
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import csv

router = APIRouter()


@router.get("/export/inventory")
def export_inventory_excel(year_month: Optional[str] = None):
    """Mar inventory 92열 양식 그대로 엑셀 다운로드
    A~S 마스터 + T~AA DC연도 + AB~AD 총입고/총출고/전월 + AE~BI 입고1~31 + BJ~CN 출고1~31
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mar inventory"

    # ─── 스타일 ───
    hdr_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ─── 헤더 구성 (92열) ───
    # 행1: 빈 행 (원본과 동일)
    # 행2: 헤더
    master_headers = [
        "Central", "Sales team", "VENDER", "SR#", "FAMILY", "DID#", "Part#",
        "MOBIS ID", "unit", "site", "MOQ", "Package", "FAB",
        "Q'ty", "SALES", "CUSTOMER", "CRD", "booking", "available Q'ty",
    ]  # A~S (19열)
    dc_headers = [f"Datecode\n{y}" for y in range(2019, 2027)]  # T~AA (8열)
    summary_headers = ["총입고", "총출고", "전월"]  # AB~AD (3열)
    in_headers = [f"입고{d}" for d in range(1, 32)]   # AE~BI (31열)
    out_headers = [f"출고{d}" for d in range(1, 32)]  # BJ~CN (31열)

    all_headers = master_headers + dc_headers + summary_headers + in_headers + out_headers
    # 총 19 + 8 + 3 + 31 + 31 = 92열

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    # ─── 데이터 ───
    # 자동 감지: year_month 없으면 가장 최근 월
    with get_db() as conn:
        if not year_month:
            r = conn.execute(
                "SELECT year_month FROM daily_inventory ORDER BY year_month DESC LIMIT 1"
            ).fetchone()
            year_month = r["year_month"] if r else ""

        masters = conn.execute(
            "SELECT * FROM product_master ORDER BY part_number"
        ).fetchall()

        for i, pm in enumerate(masters, 3):  # 데이터는 3행부터
            # A~S: 마스터
            row_data = [
                pm["central"], pm["sales_team"], pm["vender"], pm["sr_code"],
                pm["family"], pm["did"], pm["part_number"], pm["mobis_id"],
                pm["unit"], pm["site"], pm["moq"], pm["package"], pm["fab"],
                pm["current_qty"], pm["sales_person"], pm["customer"],
                pm["crd"], pm["booking"], pm["available_qty"],
            ]
            # T~AA: DC 연도별
            for y in range(2019, 2027):
                row_data.append(pm[f"dc_{y}"] or 0)
            # AB~AD: 총입고, 총출고, 전월
            row_data.append(pm["total_inbound"] or 0)
            row_data.append(pm["total_outbound"] or 0)
            row_data.append(pm["prev_month_balance"] or 0)

            # AE~BI: 입고 1~31일
            pn = pm["part_number"]
            daily = {}
            if year_month:
                daily_rows = conn.execute(
                    "SELECT day, inbound_qty, outbound_qty FROM daily_inventory WHERE part_number = ? AND year_month = ?",
                    (pn, year_month),
                ).fetchall()
                for dr in daily_rows:
                    daily[dr["day"]] = (dr["inbound_qty"] or 0, dr["outbound_qty"] or 0)

            for d in range(1, 32):
                inb, _ = daily.get(d, (0, 0))
                row_data.append(inb if inb else "")
            # BJ~CN: 출고 1~31일
            for d in range(1, 32):
                _, outb = daily.get(d, (0, 0))
                row_data.append(outb if outb else "")

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin
                if col <= 19:
                    cell.font = Font(size=9)

    # 열 너비
    widths = {1: 10, 2: 10, 3: 10, 4: 8, 5: 15, 6: 8, 7: 30, 8: 18,
              9: 5, 10: 8, 11: 8, 12: 10, 13: 5, 14: 10, 15: 8, 16: 18,
              17: 8, 18: 8, 19: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    # DC~출고 열은 좁게
    for c in range(20, 93):
        ws.column_dimensions[get_column_letter(c)].width = 7

    # 틀 고정 (G열까지 + 헤더 2행)
    ws.freeze_panes = "H3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Mar_inventory_{year_month or 'all'}.xlsx"},
    )


@router.get("/export/datecode")
def export_datecode_excel(sales_team: str = Query(...)):
    """DATECODE 엑셀 내보내기 — 영업실별"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"DATECODE({sales_team})"

    hdr_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "입고일", "SR#", "PART#", "Q'ty", "DATECODE",
        "담당 SALES", "CUSTOMER", "REMARK", "실재고",
        "출고일", "출고 CUSTOMER", "출고 PART#", "출고 Q'ty",
        "출고 SALES", "출고 REMARK", "상태",
        "외화단가(USD)", "금액(USD)", "환율", "금액(KRW)",
        "DC환산일", "경과일수(노후)", "경과일수(리드타임)", "노후도",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    with get_db() as conn:
        rows = conn.execute(
            """SELECT di.*, pm.customer as pm_customer
               FROM datecode_inventory di
               LEFT JOIN (
                   SELECT part_number, customer,
                          ROW_NUMBER() OVER (PARTITION BY part_number ORDER BY id DESC) as rn
                   FROM product_master
               ) pm ON di.part_number = pm.part_number AND pm.rn = 1
               WHERE di.sales_team = ?
               ORDER BY di.id""",
            (sales_team,),
        ).fetchall()

        for ri, r in enumerate(rows, 2):
            # 리드타임 계산
            lead_time = None
            if r["inbound_date"] and r["datecode_date"]:
                try:
                    from datetime import datetime
                    ib = datetime.strptime(r["inbound_date"][:10], "%Y-%m-%d")
                    dc = datetime.strptime(r["datecode_date"][:10], "%Y-%m-%d")
                    lead_time = (ib - dc).days
                except (ValueError, TypeError):
                    pass

            urgency_label = {"normal": "정상", "warning": "주의", "critical": "긴급"}.get(r["urgency"], "")
            vals = [
                r["inbound_date"] or "", r["sr_number"] or "", r["part_number"] or "",
                r["quantity"] or 0, r["datecode"] or "",
                r["sales_person"] or "", r["pm_customer"] or r["customer"] or "",
                r["remark"] or "", r["actual_stock"] or 0,
                r["outbound_date"] or "", r["out_customer"] or "",
                r["out_part_number"] or "", r["out_quantity"] or 0,
                r["out_sales"] or "", r["out_remark"] or "",
                r["status"] or "",
                r["unit_price_usd"] or 0, r["amount_usd"] or 0,
                r["exchange_rate"] or 0, r["amount_krw"] or 0,
                r["datecode_date"] or "", r["days_elapsed"] or 0,
                lead_time, urgency_label,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin

    # 컬럼 너비
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_team = sales_team.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=DATECODE_{safe_team}.xlsx"},
    )


@router.get("/export/shipments")
def export_shipments_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer: Optional[str] = None,
    part_number: Optional[str] = None,
):
    """출고 이력 엑셀 다운로드"""
    conditions = []
    params = []
    if start_date:
        conditions.append("ship_date >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("ship_date <= ?")
        params.append(end_date)
    if customer:
        conditions.append("customer LIKE ?")
        params.append(f"%{customer}%")
    if part_number:
        conditions.append("part_number LIKE ?")
        params.append(f"%{part_number}%")

    where = " AND ".join(conditions) if conditions else "1=1"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "shipping management"

    hdr_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["DATE", "CUSTOMER", "PART#", "Q'ty", "SALES", "lot number", "DATECODE"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM shipment_log WHERE {where} ORDER BY ship_date DESC",
            params,
        ).fetchall()
        for i, r in enumerate(rows, 2):
            values = [
                r["ship_date"], r["customer"], r["part_number"],
                r["quantity"], r["sales_person"], r["lot_number"], r["datecode"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin

    widths = {1: 12, 2: 20, 3: 30, 4: 12, 5: 10, 6: 15, 7: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"shipping_management.xlsx\""},
    )


@router.get("/export/ledger")
def export_ledger_excel():
    """수불부 엑셀 내보내기 — 상품수불명세서 형태"""
    YEAR_MONTH = "2026-03"
    YEAR_START = "2026-01"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품수불명세서"

    # 스타일
    hdr_fill_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    hdr_fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    hdr_fill_blue = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    hdr_fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    hdr_fill_gray = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    hdr_font = Font(bold=True, size=8)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 행1: 타이틀
    ws.merge_cells("A1:AG1")
    title_cell = ws.cell(row=1, column=1, value="상 품 수 불 명 세 서")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 행2: 대분류
    row2 = [
        ("DESCRIP", hdr_fill_gray), ("제조사", hdr_fill_gray), ("매입처", hdr_fill_gray),
    ]
    # 입고 16컬럼
    for _ in range(16):
        row2.append(("입고", hdr_fill_green))
    # 평균단가
    row2.append(("평균단가", hdr_fill_yellow))
    # 출고 12컬럼
    for _ in range(12):
        row2.append(("당기출고", hdr_fill_red))
    # 기말 4컬럼
    for _ in range(4):
        row2.append(("기말재고", hdr_fill_blue))

    for col, (val, fill) in enumerate(row2, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin

    # 행3: 중분류
    mid_headers = [
        "", "", "",
        "전기이월", "", "", "",
        "전월까지 누적입고", "", "", "",
        "당월입고", "", "", "",
        "(기초합산) 당기입고 계", "", "", "",
        "",  # 평균단가
        "전월까지 누적출고", "", "", "",
        "당월출고", "", "", "",
        "매출누계", "", "", "",
        "", "", "", "",
    ]
    fills3 = [hdr_fill_gray]*3 + [hdr_fill_green]*16 + [hdr_fill_yellow] + [hdr_fill_red]*12 + [hdr_fill_blue]*4
    for col, val in enumerate(mid_headers, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = fills3[col-1] if col-1 < len(fills3) else hdr_fill_gray
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin

    # 행4: 소분류 (수량/Credit(차감)/수량/금액(Credit) 반복)
    sub4 = ["수량", "Credit(차감)", "수량", "금액(Credit)"]
    detail_headers = ["DESCRIP", "제조사", "매입처"]
    for _ in range(4):  # 입고 4그룹
        detail_headers += sub4
    detail_headers.append("Credit(차감)")  # 평균단가
    for _ in range(3):  # 출고 3그룹
        detail_headers += sub4
    detail_headers += sub4  # 기말
    fills4 = [hdr_fill_gray]*3 + [hdr_fill_green]*16 + [hdr_fill_yellow] + [hdr_fill_red]*12 + [hdr_fill_blue]*4
    for col, val in enumerate(detail_headers, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.fill = fills4[col-1] if col-1 < len(fills4) else hdr_fill_gray
        cell.font = Font(size=8)
        cell.alignment = center
        cell.border = thin

    # 데이터
    with get_db() as conn:
        pm_join = """LEFT JOIN (
            SELECT part_number, family, vender, customer,
                   ROW_NUMBER() OVER (PARTITION BY part_number ORDER BY id DESC) as rn
            FROM product_master
        ) pm ON di.part_number = pm.part_number AND pm.rn = 1"""

        rows = conn.execute(f"""
            SELECT di.part_number,
                   COALESCE(pm.family, '') as family,
                   COALESCE(pm.vender, '') as vender,
                   COALESCE(pm.customer, '') as customer,
                   SUM(CASE WHEN di.inbound_date = '' OR di.inbound_date IS NULL OR SUBSTR(di.inbound_date,1,7) < ?
                       THEN di.quantity ELSE 0 END) as carry_forward,
                   SUM(CASE WHEN di.inbound_date = '' OR di.inbound_date IS NULL OR SUBSTR(di.inbound_date,1,7) < ?
                       THEN di.amount_krw ELSE 0 END) as carry_forward_krw,
                   SUM(CASE WHEN SUBSTR(di.inbound_date,1,7) >= ? AND SUBSTR(di.inbound_date,1,7) < ?
                       THEN di.quantity ELSE 0 END) as cum_in_prev,
                   SUM(CASE WHEN SUBSTR(di.inbound_date,1,7) >= ? AND SUBSTR(di.inbound_date,1,7) < ?
                       THEN di.amount_krw ELSE 0 END) as cum_in_prev_krw,
                   SUM(CASE WHEN SUBSTR(di.inbound_date,1,7) = ?
                       THEN di.quantity ELSE 0 END) as cur_in,
                   SUM(CASE WHEN SUBSTR(di.inbound_date,1,7) = ?
                       THEN di.amount_krw ELSE 0 END) as cur_in_krw,
                   SUM(CASE WHEN di.out_quantity > 0 AND SUBSTR(di.outbound_date,1,7) != ?
                       THEN di.out_quantity ELSE 0 END) as cum_out_prev,
                   SUM(CASE WHEN di.out_quantity > 0 AND SUBSTR(di.outbound_date,1,7) != ?
                       THEN di.out_quantity * di.unit_price_usd * di.exchange_rate ELSE 0 END) as cum_out_prev_krw,
                   SUM(CASE WHEN di.out_quantity > 0 AND SUBSTR(di.outbound_date,1,7) = ?
                       THEN di.out_quantity ELSE 0 END) as cur_out,
                   SUM(CASE WHEN di.out_quantity > 0 AND SUBSTR(di.outbound_date,1,7) = ?
                       THEN di.out_quantity * di.unit_price_usd * di.exchange_rate ELSE 0 END) as cur_out_krw,
                   SUM(di.actual_stock) as end_balance,
                   SUM(CASE WHEN di.actual_stock > 0 THEN di.actual_stock * di.unit_price_usd * di.exchange_rate ELSE 0 END) as end_krw,
                   CASE WHEN SUM(di.quantity) > 0 THEN SUM(di.amount_krw) / SUM(di.quantity) ELSE 0 END as avg_price,
                   COALESCE(MAX(cp.credit_usd), 0) as credit_usd
            FROM datecode_inventory di
            {pm_join}
            LEFT JOIN credit_price cp ON di.part_number = cp.part_number
            GROUP BY di.part_number
            ORDER BY SUM(di.actual_stock) DESC
        """, [YEAR_START, YEAR_START, YEAR_START, YEAR_MONTH, YEAR_START, YEAR_MONTH,
              YEAR_MONTH, YEAR_MONTH, YEAR_MONTH, YEAR_MONTH, YEAR_MONTH, YEAR_MONTH]).fetchall()

        for i, r in enumerate(rows, 5):
            cf = r["carry_forward"] or 0
            cf_krw = round(r["carry_forward_krw"] or 0)
            cip = r["cum_in_prev"] or 0
            cip_krw = round(r["cum_in_prev_krw"] or 0)
            ci = r["cur_in"] or 0
            ci_krw = round(r["cur_in_krw"] or 0)
            cit = cip + ci
            cit_krw = cip_krw + ci_krw
            ig = cf + cit
            ig_krw = cf_krw + cit_krw
            cop = r["cum_out_prev"] or 0
            cop_krw = round(r["cum_out_prev_krw"] or 0)
            co = r["cur_out"] or 0
            co_krw = round(r["cur_out_krw"] or 0)
            cot = cop + co
            cot_krw = cop_krw + co_krw
            eb = r["end_balance"] or 0
            eb_krw = round(r["end_krw"] or 0)
            cusd = r["credit_usd"] or 0

            row_data = [
                r["part_number"], r["vender"], r["customer"],
                # 전기이월
                cf, round(cf * cusd) if cusd > 0 else "", cf, cf_krw,
                # 전월누적입고
                cip, round(cip * cusd) if cusd > 0 else "", cip, cip_krw,
                # 당월입고
                ci, round(ci * cusd) if cusd > 0 else "", ci, ci_krw,
                # 당기입고계
                ig, round(ig * cusd) if cusd > 0 else "", ig, ig_krw,
                # 평균단가
                round(r["avg_price"] or 0),
                # 전월누적출고
                cop, round(cop * cusd) if cusd > 0 else "", cop, cop_krw,
                # 당월출고
                co, round(co * cusd) if cusd > 0 else "", co, co_krw,
                # 매출누계
                cot, round(cot * cusd) if cusd > 0 else "", cot, cot_krw,
                # 기말재고
                eb, round(eb * cusd) if cusd > 0 else "", eb, eb_krw,
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val if val != "" else None)
                cell.border = thin
                cell.font = Font(size=8)
                if isinstance(val, (int, float)) and val != "":
                    cell.number_format = '#,##0'

    # 열 너비
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    for c in range(4, 37):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = "D5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"ledger_export.xlsx\""},
    )
