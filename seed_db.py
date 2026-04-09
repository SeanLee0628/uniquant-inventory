"""Mar_Inv.xlsx → backend/inventory.db 시드 스크립트"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))

# DB_PATH를 로컬 backend/inventory.db로 고정
os.environ["DB_PATH"] = os.path.join(os.path.dirname(__file__), "backend", "inventory.db")

import sqlite3
import openpyxl
from datetime import date, datetime, timedelta
import re

DB_PATH = os.environ["DB_PATH"]
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else "C:/Users/user/Downloads/Mar_Inv.xlsx"

# ── 유틸 ──

def safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s == "." else s

def safe_int(val):
    if val is None:
        return 0
    s = str(val).strip()
    if s in ("", ".", "None", "N/A", "-"):
        return 0
    try:
        return int(float(s.replace(",", "")))
    except (ValueError, TypeError):
        return 0

def safe_float(val):
    if val is None:
        return 0.0
    s = str(val).strip()
    if s in ("", ".", "None", "N/A", "-"):
        return 0.0
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return 0.0

def safe_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()

def datecode_to_date(dc):
    if not dc or not re.match(r"^\d{6}$", str(dc).strip()):
        return None
    s = str(dc).strip()
    year = int(s[:4])
    week = int(s[4:6])
    if week < 1 or week > 53 or year < 2000 or year > 2100:
        return None
    return date(year, 1, 1) + timedelta(days=(week - 1) * 7)

def calc_urgency(days_elapsed):
    if days_elapsed < 365:
        return "normal"
    elif days_elapsed <= 730:
        return "warning"
    return "critical"


# ── DB 초기화 ──

from database import init_db
init_db()

print(f"DB: {DB_PATH}")
print(f"Excel: {EXCEL_PATH}")

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
conn.execute("PRAGMA journal_mode=WAL")
conn.execute("PRAGMA foreign_keys=ON")

# ═══ 1. Mar inventory ═══
ws = None
for name in wb.sheetnames:
    if "mar" in name.lower() or "inventory" in name.lower():
        ws = wb[name]
        break
if ws is None:
    ws = wb[wb.sheetnames[0]]

rows = list(ws.iter_rows(values_only=True))
data_rows = rows[2:]  # 헤더 2행 스킵

year_month = "2026-03"  # Mar inventory

conn.execute("DELETE FROM product_master")
conn.execute("DELETE FROM daily_inventory")
conn.execute("DELETE FROM monthly_ledger")

inserted = 0
for row in data_rows:
    pn = safe_str(row[6]) if len(row) > 6 else ""
    if not pn:
        continue

    current_qty = safe_int(row[13]) if len(row) > 13 else 0
    booking = safe_int(row[17]) if len(row) > 17 else 0
    available = safe_int(row[18]) if len(row) > 18 else (current_qty - booking)

    params = (
        safe_str(row[0]), safe_str(row[1]), safe_str(row[2]), safe_str(row[3]),
        safe_str(row[4]), safe_str(row[5]), pn,
        safe_str(row[7]) if len(row) > 7 else "",
        safe_str(row[8]) if len(row) > 8 else "EA",
        safe_str(row[9]) if len(row) > 9 else "",
        safe_int(row[10]) if len(row) > 10 else 0,
        safe_str(row[11]) if len(row) > 11 else "",
        safe_str(row[12]) if len(row) > 12 else "",
        current_qty,
        safe_str(row[14]) if len(row) > 14 else "",
        safe_str(row[15]) if len(row) > 15 else "",
        safe_str(row[16]) if len(row) > 16 else "",
        booking, available,
        safe_int(row[19]) if len(row) > 19 else 0,
        safe_int(row[20]) if len(row) > 20 else 0,
        safe_int(row[21]) if len(row) > 21 else 0,
        safe_int(row[22]) if len(row) > 22 else 0,
        safe_int(row[23]) if len(row) > 23 else 0,
        safe_int(row[24]) if len(row) > 24 else 0,
        safe_int(row[25]) if len(row) > 25 else 0,
        safe_int(row[26]) if len(row) > 26 else 0,
        safe_int(row[27]) if len(row) > 27 else 0,
        safe_int(row[28]) if len(row) > 28 else 0,
        safe_int(row[29]) if len(row) > 29 else 0,
    )

    conn.execute("""
        INSERT INTO product_master (
            central, sales_team, vender, sr_code, family, did,
            part_number, mobis_id, unit, site, moq, package, fab,
            current_qty, sales_person, customer, crd, booking, available_qty,
            dc_2019, dc_2020, dc_2021, dc_2022, dc_2023,
            dc_2024, dc_2025, dc_2026,
            total_inbound, total_outbound, prev_month_balance
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, params)
    inserted += 1

    # 일별 입출고
    if len(row) > 61:
        for day in range(1, 32):
            in_col = 29 + day
            out_col = 60 + day
            in_qty = safe_int(row[in_col]) if len(row) > in_col else 0
            out_qty = safe_int(row[out_col]) if len(row) > out_col else 0
            if in_qty > 0 or out_qty > 0:
                conn.execute(
                    """INSERT INTO daily_inventory (part_number, year_month, day, inbound_qty, outbound_qty)
                       VALUES (?, ?, ?, ?, ?)
                       ON CONFLICT(part_number, year_month, day)
                       DO UPDATE SET inbound_qty = excluded.inbound_qty,
                                    outbound_qty = excluded.outbound_qty""",
                    (pn, year_month, day, in_qty, out_qty),
                )

# monthly_ledger
ledger_count = 0
for row in data_rows:
    pn = safe_str(row[6]) if len(row) > 6 else ""
    if not pn:
        continue
    prev = safe_int(row[29]) if len(row) > 29 else 0
    cur = safe_int(row[13]) if len(row) > 13 else 0
    bk = safe_int(row[17]) if len(row) > 17 else 0
    avail = safe_int(row[18]) if len(row) > 18 else (cur - bk)
    m_in = sum(safe_int(row[c]) for c in range(30, min(61, len(row))))
    m_out = sum(safe_int(row[c]) for c in range(61, min(92, len(row))))

    if prev > 0 or m_in > 0 or m_out > 0 or cur > 0:
        conn.execute(
            """INSERT INTO monthly_ledger
               (year_month, part_number, family, vender, customer,
                prev_balance, month_inbound, month_outbound, end_balance,
                booking, available_qty)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(year_month, part_number)
               DO UPDATE SET prev_balance=excluded.prev_balance,
                             month_inbound=excluded.month_inbound,
                             month_outbound=excluded.month_outbound,
                             end_balance=excluded.end_balance,
                             booking=excluded.booking,
                             available_qty=excluded.available_qty""",
            (year_month, pn,
             safe_str(row[4]) if len(row) > 4 else "",
             safe_str(row[2]) if len(row) > 2 else "",
             safe_str(row[15]) if len(row) > 15 else "",
             prev, m_in, m_out, cur, bk, avail),
        )
        ledger_count += 1

print(f"[Mar inventory] inserted: {inserted}, ledger: {ledger_count}")

# ═══ 2. Shipping management ═══
ws_ship = None
for name in wb.sheetnames:
    if "shipping" in name.lower():
        ws_ship = wb[name]
        break

if ws_ship:
    rows_ship = list(ws_ship.iter_rows(values_only=True))
    conn.execute("DELETE FROM shipment_log")
    ship_count = 0
    for row in rows_ship[1:]:
        pn = safe_str(row[2]) if len(row) > 2 else ""
        qty = safe_int(row[3]) if len(row) > 3 else 0
        if not pn or qty <= 0:
            continue
        ship_date = safe_date(row[0]) if row[0] else ""
        customer = safe_str(row[1]) if len(row) > 1 else ""
        sales = safe_str(row[4]) if len(row) > 4 else ""
        lot = safe_str(row[5]) if len(row) > 5 else ""
        dc = safe_str(row[6]) if len(row) > 6 else ""
        if lot == ".": lot = ""
        if dc == ".": dc = ""

        conn.execute(
            """INSERT INTO shipment_log
               (ship_date, customer, part_number, quantity, sales_person, lot_number, datecode)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (ship_date, customer, pn, qty, sales, lot, dc),
        )
        ship_count += 1

        if ship_date:
            try:
                d = datetime.strptime(ship_date, "%Y-%m-%d")
                ym = d.strftime("%Y-%m")
                day_num = d.day
                conn.execute(
                    """INSERT INTO daily_inventory (part_number, year_month, day, outbound_qty)
                       VALUES (?, ?, ?, ?)
                       ON CONFLICT(part_number, year_month, day)
                       DO UPDATE SET outbound_qty = outbound_qty + excluded.outbound_qty""",
                    (pn, ym, day_num, qty),
                )
            except (ValueError, TypeError):
                pass

    print(f"[Shipping] inserted: {ship_count}")

# ═══ 3. DATECODE 시트 ═══

def detect_sales_team(sheet_name, headers):
    name = sheet_name.replace(" ", "")
    if "1" in name and ("영업" in name or "실" in name or "1실" in name):
        return "영업1실"
    if "2" in name and ("영업" in name or "실" in name or "2실" in name):
        return "영업2실"
    if len(headers) >= 8:
        f_header = str(headers[5]).strip().upper() if headers[5] else ""
        if "PO" in f_header:
            return "영업2실"
        if "SALES" in f_header or "담당" in f_header:
            return "영업1실"
    return None

conn.execute("DELETE FROM datecode_inventory")

for sheet_name in wb.sheetnames:
    if "datecode" not in sheet_name.lower():
        continue
    ws_dc = wb[sheet_name]
    rows_dc = list(ws_dc.iter_rows(values_only=True))
    if len(rows_dc) < 2:
        continue

    headers = [safe_str(h) for h in rows_dc[0]]
    team = detect_sales_team(sheet_name, headers)
    if team is None:
        # Fallback: sheet name contains 1 or 2
        if "1" in sheet_name:
            team = "영업1실"
        elif "2" in sheet_name:
            team = "영업2실"
        else:
            continue

    dc_count = 0
    for row in rows_dc[1:]:
        if len(row) < 8:
            continue
        pn = safe_str(row[2])
        if not pn:
            continue

        dc_str = safe_str(row[4])
        dc_date = datecode_to_date(dc_str)
        days_elapsed = (date.today() - dc_date).days if dc_date else 0
        qty = safe_int(row[3])

        if team == "영업1실":
            status = safe_str(row[15]) if len(row) > 15 else "사용가능"
            if not status: status = "사용가능"
            out_qty = safe_int(row[12]) if len(row) > 12 else 0
            actual = (qty - out_qty) if status == "완료" else qty
            unit_usd = safe_float(row[16]) if len(row) > 16 else 0
            exrate = safe_float(row[18]) if len(row) > 18 else 0

            parsed = {
                "sales_team": "영업1실",
                "inbound_date": safe_date(row[0]),
                "sr_number": safe_str(row[1]),
                "part_number": pn,
                "quantity": qty,
                "datecode": dc_str,
                "datecode_date": dc_date.isoformat() if dc_date else "",
                "days_elapsed": days_elapsed,
                "sales_person": safe_str(row[5]),
                "customer": safe_str(row[6]),
                "po_number": "",
                "remark": safe_str(row[7]),
                "actual_stock": actual,
                "outbound_date": safe_date(row[9]) if len(row) > 9 else "",
                "out_customer": safe_str(row[10]) if len(row) > 10 else "",
                "out_part_number": safe_str(row[11]) if len(row) > 11 else "",
                "out_quantity": out_qty,
                "out_sales": safe_str(row[13]) if len(row) > 13 else "",
                "out_remark": safe_str(row[14]) if len(row) > 14 else "",
                "status": status,
                "unit_price_usd": unit_usd,
                "amount_usd": actual * unit_usd,
                "exchange_rate": exrate,
                "amount_krw": actual * unit_usd * exrate,
                "urgency": calc_urgency(days_elapsed),
            }
        else:
            status = safe_str(row[14]) if len(row) > 14 else "사용가능"
            if not status: status = "사용가능"
            out_qty = safe_int(row[11]) if len(row) > 11 else 0
            actual = safe_int(row[7]) if row[7] is not None else qty
            unit_usd = safe_float(row[15]) if len(row) > 15 else 0
            exrate = safe_float(row[17]) if len(row) > 17 else 0

            parsed = {
                "sales_team": "영업2실",
                "inbound_date": safe_date(row[0]),
                "sr_number": safe_str(row[1]),
                "part_number": pn,
                "quantity": qty,
                "datecode": dc_str,
                "datecode_date": dc_date.isoformat() if dc_date else "",
                "days_elapsed": days_elapsed,
                "sales_person": "",
                "customer": "",
                "po_number": safe_str(row[5]),
                "remark": safe_str(row[6]),
                "actual_stock": actual,
                "outbound_date": safe_date(row[8]) if len(row) > 8 else "",
                "out_customer": safe_str(row[9]) if len(row) > 9 else "",
                "out_part_number": safe_str(row[10]) if len(row) > 10 else "",
                "out_quantity": out_qty,
                "out_sales": safe_str(row[12]) if len(row) > 12 else "",
                "out_remark": safe_str(row[13]) if len(row) > 13 else "",
                "status": status,
                "unit_price_usd": unit_usd,
                "amount_usd": actual * unit_usd,
                "exchange_rate": exrate,
                "amount_krw": actual * unit_usd * exrate,
                "urgency": calc_urgency(days_elapsed),
            }

        conn.execute("""
            INSERT INTO datecode_inventory (
                sales_team, inbound_date, sr_number, part_number, quantity,
                datecode, datecode_date, days_elapsed, sales_person, customer,
                po_number, remark, actual_stock, outbound_date, out_customer,
                out_part_number, out_quantity, out_sales, out_remark, status,
                unit_price_usd, amount_usd, exchange_rate, amount_krw, urgency
            ) VALUES (
                :sales_team, :inbound_date, :sr_number, :part_number, :quantity,
                :datecode, :datecode_date, :days_elapsed, :sales_person, :customer,
                :po_number, :remark, :actual_stock, :outbound_date, :out_customer,
                :out_part_number, :out_quantity, :out_sales, :out_remark, :status,
                :unit_price_usd, :amount_usd, :exchange_rate, :amount_krw, :urgency
            )
        """, parsed)
        dc_count += 1

    print(f"[DATECODE {team}] inserted: {dc_count}")

wb.close()
conn.commit()

# 최종 확인
for tbl in ["product_master", "datecode_inventory", "shipment_log", "daily_inventory", "monthly_ledger"]:
    cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
    print(f"  {tbl}: {cnt} rows")

db_size = os.path.getsize(DB_PATH)
print(f"\nDB size: {db_size / 1024 / 1024:.1f} MB")
conn.close()
print("Done!")
