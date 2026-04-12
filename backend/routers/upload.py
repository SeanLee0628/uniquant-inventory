"""엑셀 업로드 — Mar inventory(품목 마스터) + DATECODE(거래 원장)"""

from fastapi import APIRouter, UploadFile, File, Query
from datetime import date, timedelta, datetime
from database import get_db
import openpyxl
import io
import re
import json

router = APIRouter()


# ─── 유틸 ───

def safe_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s == ".":
        return ""
    return s


def safe_int(val) -> int:
    if val is None:
        return 0
    s = str(val).strip()
    if s in ("", ".", "None", "N/A", "-"):
        return 0
    try:
        return int(float(s.replace(",", "")))
    except (ValueError, TypeError):
        return 0


def safe_float(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if s in ("", ".", "None", "N/A", "-"):
        return 0.0
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def safe_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def datecode_to_date(dc: str) -> date | None:
    if not dc or not re.match(r"^\d{6}$", str(dc).strip()):
        return None
    s = str(dc).strip()
    year = int(s[:4])
    week = int(s[4:6])
    if week < 1 or week > 53 or year < 2000 or year > 2100:
        return None
    return date(year, 1, 1) + timedelta(days=(week - 1) * 7)


def calc_urgency(days_elapsed: int) -> str:
    if days_elapsed < 365:
        return "normal"
    elif days_elapsed <= 730:
        return "warning"
    else:
        return "critical"


# ═══════════════════════════════════════════════════
# ① Mar inventory 업로드 (품목 마스터)
# ═══════════════════════════════════════════════════

@router.post("/upload/master")
async def upload_master(
    file: UploadFile = File(...),
    year_month: str = Query("", description="일별 입출고 기준 월 (YYYY-MM). 예: 2026-03"),
):
    """Mar inventory 엑셀 업로드 — 헤더 2행, 데이터 3행부터. Part# UPSERT. AE~CN 일별 입출고 포함."""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Mar inventory 시트 찾기
    ws = None
    for name in wb.sheetnames:
        if "mar" in name.lower() or "inventory" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return {"total": 0, "inserted": 0, "updated": 0, "errors": 0}

    # year_month 자동 추출: 파라미터 없으면 파일명에서 시도, 그래도 없으면 이번 달
    if not year_month:
        # 파일명에서 월 추출 시도 (예: 자재_3월.xlsx → 올해-03)
        fname = file.filename or ""
        for m in range(1, 13):
            if f"{m}월" in fname:
                year_month = f"{date.today().year}-{m:02d}"
                break
        if not year_month:
            year_month = date.today().strftime("%Y-%m")

    data_rows = rows[2:]

    total = 0
    inserted = 0
    updated = 0
    errors = 0
    has_stock = 0
    daily_count = 0

    with get_db() as conn:
        # 기존 데이터 삭제 후 전체 재삽입 (원본 행 구조 유지)
        conn.execute("DELETE FROM product_master")

        for row in data_rows:
            total += 1
            pn = safe_str(row[6]) if len(row) > 6 else ""  # G: Part#
            if not pn:
                errors += 1
                continue

            current_qty = safe_int(row[13]) if len(row) > 13 else 0  # N: Q'ty
            booking = safe_int(row[17]) if len(row) > 17 else 0      # R: booking
            available = safe_int(row[18]) if len(row) > 18 else (current_qty - booking)  # S: available

            if current_qty > 0:
                has_stock += 1

            params = (
                safe_str(row[0]),   # A: Central
                safe_str(row[1]),   # B: Sales team
                safe_str(row[2]),   # C: VENDER
                safe_str(row[3]),   # D: SR#
                safe_str(row[4]),   # E: FAMILY
                safe_str(row[5]),   # F: DID#
                pn,                 # G: Part#
                safe_str(row[7]) if len(row) > 7 else "",   # H: MOBIS ID
                safe_str(row[8]) if len(row) > 8 else "EA", # I: unit
                safe_str(row[9]) if len(row) > 9 else "",   # J: site
                safe_int(row[10]) if len(row) > 10 else 0,  # K: MOQ
                safe_str(row[11]) if len(row) > 11 else "",  # L: Package
                safe_str(row[12]) if len(row) > 12 else "",  # M: FAB
                current_qty,         # N: Q'ty (현재고)
                safe_str(row[14]) if len(row) > 14 else "",  # O: SALES
                safe_str(row[15]) if len(row) > 15 else "",  # P: CUSTOMER
                safe_str(row[16]) if len(row) > 16 else "",  # Q: CRD
                booking,             # R: booking
                available,           # S: available Q'ty
                safe_int(row[19]) if len(row) > 19 else 0,  # T: DC 2019
                safe_int(row[20]) if len(row) > 20 else 0,  # U: DC 2020
                safe_int(row[21]) if len(row) > 21 else 0,  # V: DC 2021
                safe_int(row[22]) if len(row) > 22 else 0,  # W: DC 2022
                safe_int(row[23]) if len(row) > 23 else 0,  # X: DC 2023
                safe_int(row[24]) if len(row) > 24 else 0,  # Y: DC 2024
                safe_int(row[25]) if len(row) > 25 else 0,  # Z: DC 2025
                safe_int(row[26]) if len(row) > 26 else 0,  # AA: DC 2026
                safe_int(row[27]) if len(row) > 27 else 0,  # AB: 총입고
                safe_int(row[28]) if len(row) > 28 else 0,  # AC: 총출고
                safe_int(row[29]) if len(row) > 29 else 0,  # AD: 전월
            )

            conn.execute("""
                INSERT INTO product_master (
                    central, sales_team, vender, sr_code, family, did,
                    part_number, mobis_id, unit, site, moq, package, fab,
                    current_qty, sales_person, customer, crd, booking, available_qty,
                    dc_2019, dc_2020, dc_2021, dc_2022, dc_2023,
                    dc_2024, dc_2025, dc_2026,
                    total_inbound, total_outbound, prev_month_balance
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, params)
            inserted += 1

        # AE~CN 일별 입출고를 daily_inventory에 저장
        for row in data_rows:
            pn = safe_str(row[6]) if len(row) > 6 else ""
            if not pn or len(row) <= 61:
                continue
            for day in range(1, 32):
                in_col = 29 + day   # AE=30 → day1, AF=31 → day2, ... BI=60 → day31
                out_col = 60 + day  # BJ=61 → day1, BK=62 → day2, ... CN=91 → day31
                in_qty = safe_int(row[in_col]) if len(row) > in_col else 0
                out_qty = safe_int(row[out_col]) if len(row) > out_col else 0
                if in_qty > 0 or out_qty > 0:
                    conn.execute(
                        """INSERT INTO daily_inventory (part_number, year_month, day, inbound_qty, outbound_qty)
                           VALUES (?, ?, ?, ?, ?)
                           ON CONFLICT(part_number, year_month, day)
                           DO UPDATE SET inbound_qty = excluded.inbound_qty,
                                        outbound_qty = excluded.outbound_qty""",
                        (pn, year_month, day, in_qty, out_qty),
                    )
                    daily_count += 1

        # monthly_ledger에 전기이월/당기입고/당기출고/기말재고 저장
        conn.execute("DELETE FROM monthly_ledger WHERE year_month = ?", (year_month,))
        ledger_count = 0
        for row in data_rows:
            pn = safe_str(row[6]) if len(row) > 6 else ""
            if not pn:
                continue
            prev = safe_int(row[29]) if len(row) > 29 else 0   # AD: 전월잔고
            cur = safe_int(row[13]) if len(row) > 13 else 0    # N: 현재고
            bk = safe_int(row[17]) if len(row) > 17 else 0     # R: booking
            avail = safe_int(row[18]) if len(row) > 18 else (cur - bk)

            # 당월 입고/출고 합 (일별 데이터에서)
            m_in = sum(safe_int(row[c]) for c in range(30, min(61, len(row))))
            m_out = sum(safe_int(row[c]) for c in range(61, min(92, len(row))))

            if prev > 0 or m_in > 0 or m_out > 0 or cur > 0:
                conn.execute(
                    """INSERT INTO monthly_ledger
                       (year_month, part_number, family, vender, customer,
                        prev_balance, month_inbound, month_outbound, end_balance,
                        booking, available_qty)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                       ON CONFLICT(year_month, part_number)
                       DO UPDATE SET prev_balance=excluded.prev_balance,
                                     month_inbound=excluded.month_inbound,
                                     month_outbound=excluded.month_outbound,
                                     end_balance=excluded.end_balance,
                                     booking=excluded.booking,
                                     available_qty=excluded.available_qty""",
                    (year_month, pn,
                     safe_str(row[4]) if len(row) > 4 else "",
                     safe_str(row[2]) if len(row) > 2 else "",
                     safe_str(row[15]) if len(row) > 15 else "",
                     prev, m_in, m_out, cur, bk, avail),
                )
                ledger_count += 1

    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "errors": errors,
        "has_stock": has_stock,
        "daily_imported": daily_count,
        "ledger_imported": ledger_count,
        "year_month": year_month,
    }


# ═══════════════════════════════════════════════════
# ② DATECODE 업로드 (거래 원장)
# ═══════════════════════════════════════════════════

def detect_sales_team(sheet_name: str, headers: list[str]) -> str | None:
    name = sheet_name.replace(" ", "")
    if "영업1" in name or "1실" in name:
        return "영업1실"
    if "영업2" in name or "2실" in name:
        return "영업2실"
    if len(headers) >= 8:
        f_header = str(headers[5]).strip().upper() if headers[5] else ""
        if "PO" in f_header:
            return "영업2실"
        if "SALES" in f_header or "담당" in f_header:
            return "영업1실"
    return None


def parse_team1_row(row: tuple) -> dict | None:
    """영업1실: A~V"""
    if len(row) < 9:
        return None
    pn = safe_str(row[2])
    if not pn:
        return None
    dc_str = safe_str(row[4])
    dc_date = datecode_to_date(dc_str)
    days_elapsed = (date.today() - dc_date).days if dc_date else 0

    qty = safe_int(row[3])
    status = safe_str(row[15]) if len(row) > 15 else "사용가능"
    if not status:
        status = "사용가능"
    out_qty = safe_int(row[12]) if len(row) > 12 else 0
    actual = (qty - out_qty) if status == "완료" else qty

    unit_usd = safe_float(row[16]) if len(row) > 16 else 0
    exrate = safe_float(row[18]) if len(row) > 18 else 0
    amt_usd = actual * unit_usd
    amt_krw = amt_usd * exrate

    return {
        "sales_team": "영업1실",
        "inbound_date": safe_date(row[0]),
        "sr_number": safe_str(row[1]),
        "part_number": pn,
        "quantity": qty,
        "datecode": dc_str,
        "datecode_date": dc_date.isoformat() if dc_date else "",
        "days_elapsed": days_elapsed,
        "sales_person": safe_str(row[5]),
        "customer": safe_str(row[6]),
        "po_number": "",
        "remark": safe_str(row[7]),
        "actual_stock": actual,
        "outbound_date": safe_date(row[9]) if len(row) > 9 else "",
        "out_customer": safe_str(row[10]) if len(row) > 10 else "",
        "out_part_number": safe_str(row[11]) if len(row) > 11 else "",
        "out_quantity": out_qty,
        "out_sales": safe_str(row[13]) if len(row) > 13 else "",
        "out_remark": safe_str(row[14]) if len(row) > 14 else "",
        "status": status,
        "unit_price_usd": unit_usd,
        "amount_usd": amt_usd,
        "exchange_rate": exrate,
        "amount_krw": amt_krw,
        "urgency": calc_urgency(days_elapsed),
    }


def parse_team2_row(row: tuple) -> dict | None:
    """영업2실: F=PO#, G=REMARK, H=실재고, 이후 한 칸 앞"""
    if len(row) < 8:
        return None
    pn = safe_str(row[2])
    if not pn:
        return None
    dc_str = safe_str(row[4])
    dc_date = datecode_to_date(dc_str)
    days_elapsed = (date.today() - dc_date).days if dc_date else 0

    qty = safe_int(row[3])
    status = safe_str(row[14]) if len(row) > 14 else "사용가능"
    if not status:
        status = "사용가능"
    out_qty = safe_int(row[11]) if len(row) > 11 else 0
    actual = safe_int(row[7]) if row[7] is not None else qty

    unit_usd = safe_float(row[15]) if len(row) > 15 else 0
    exrate = safe_float(row[17]) if len(row) > 17 else 0
    amt_usd = actual * unit_usd
    amt_krw = amt_usd * exrate

    return {
        "sales_team": "영업2실",
        "inbound_date": safe_date(row[0]),
        "sr_number": safe_str(row[1]),
        "part_number": pn,
        "quantity": qty,
        "datecode": dc_str,
        "datecode_date": dc_date.isoformat() if dc_date else "",
        "days_elapsed": days_elapsed,
        "sales_person": "",
        "customer": "",
        "po_number": safe_str(row[5]),
        "remark": safe_str(row[6]),
        "actual_stock": actual,
        "outbound_date": safe_date(row[8]) if len(row) > 8 else "",
        "out_customer": safe_str(row[9]) if len(row) > 9 else "",
        "out_part_number": safe_str(row[10]) if len(row) > 10 else "",
        "out_quantity": out_qty,
        "out_sales": safe_str(row[12]) if len(row) > 12 else "",
        "out_remark": safe_str(row[13]) if len(row) > 13 else "",
        "status": status,
        "unit_price_usd": unit_usd,
        "amount_usd": amt_usd,
        "exchange_rate": exrate,
        "amount_krw": amt_krw,
        "urgency": calc_urgency(days_elapsed),
    }


INSERT_DC_SQL = """
    INSERT INTO datecode_inventory (
        sales_team, inbound_date, sr_number, part_number, quantity,
        datecode, datecode_date, days_elapsed, sales_person, customer,
        po_number, remark, actual_stock, outbound_date, out_customer,
        out_part_number, out_quantity, out_sales, out_remark, status,
        unit_price_usd, amount_usd, exchange_rate, amount_krw, urgency
    ) VALUES (
        :sales_team, :inbound_date, :sr_number, :part_number, :quantity,
        :datecode, :datecode_date, :days_elapsed, :sales_person, :customer,
        :po_number, :remark, :actual_stock, :outbound_date, :out_customer,
        :out_part_number, :out_quantity, :out_sales, :out_remark, :status,
        :unit_price_usd, :amount_usd, :exchange_rate, :amount_krw, :urgency
    )
"""


@router.post("/upload/datecode")
async def upload_datecode(
    file: UploadFile = File(...),
    overwrite: bool = Query(False),
):
    """DATECODE 엑셀 업로드 — 영업1실/영업2실 자동 감지"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    results = []

    with get_db() as conn:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            headers = [safe_str(h) for h in rows[0]]
            team = detect_sales_team(sheet_name, headers)
            if team is None:
                continue

            if overwrite:
                conn.execute("DELETE FROM datecode_inventory WHERE sales_team = ?", (team,))

            parser = parse_team1_row if team == "영업1실" else parse_team2_row
            result = {"sales_team": team, "total": 0, "available": 0, "completed": 0, "waiting": 0, "errors": 0, "critical": 0}

            for row in rows[1:]:
                result["total"] += 1
                parsed = parser(row)
                if parsed is None:
                    result["errors"] += 1
                    continue
                conn.execute(INSERT_DC_SQL, parsed)
                if parsed["status"] == "완료":
                    result["completed"] += 1
                elif parsed["status"] == "대기":
                    result["waiting"] += 1
                else:
                    result["available"] += 1
                if parsed["urgency"] == "critical":
                    result["critical"] += 1

            # daily_inventory에 입고 반영
            _sync_daily_inbound(conn, team)

            results.append(result)

    wb.close()
    return results


def _sync_daily_inbound(conn, sales_team: str):
    rows = conn.execute(
        """SELECT part_number, inbound_date, quantity
           FROM datecode_inventory
           WHERE sales_team = ? AND inbound_date != ''""",
        (sales_team,),
    ).fetchall()
    for r in rows:
        try:
            d = datetime.strptime(r["inbound_date"], "%Y-%m-%d")
        except (ValueError, TypeError):
            continue
        ym = d.strftime("%Y-%m")
        day = d.day
        conn.execute(
            """INSERT INTO daily_inventory (part_number, year_month, day, inbound_qty)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(part_number, year_month, day)
               DO UPDATE SET inbound_qty = daily_inventory.inbound_qty + excluded.inbound_qty""",
            (r["part_number"], ym, day, r["quantity"]),
        )


@router.get("/upload/check-existing")
def check_existing():
    with get_db() as conn:
        row = conn.execute("SELECT COUNT(*) as cnt FROM datecode_inventory").fetchone()
        return {"exists": row["cnt"] > 0, "count": row["cnt"]}


# ═══════════════════════════════════════════════════
# ③ shipping management 일괄 임포트
# ═══════════════════════════════════════════════════

@router.post("/upload/shipping")
async def upload_shipping(
    file: UploadFile = File(...),
    overwrite: bool = Query(False),
):
    """shipping management 엑셀 일괄 임포트 — A:DATE B:CUSTOMER C:PART# D:Q'ty E:SALES F:lot G:DATECODE"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # shipping management 시트 찾기
    ws = None
    for name in wb.sheetnames:
        if "shipping" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return {"total": 0, "inserted": 0, "errors": 0}

    total = 0
    inserted = 0
    errors = 0

    with get_db() as conn:
        if overwrite:
            conn.execute("DELETE FROM shipment_log")
            conn.execute("DELETE FROM daily_inventory WHERE outbound_qty > 0")

        for row in rows[1:]:
            total += 1
            pn = safe_str(row[2]) if len(row) > 2 else ""
            qty = safe_int(row[3]) if len(row) > 3 else 0
            if not pn or qty <= 0:
                errors += 1
                continue

            ship_date = safe_date(row[0]) if row[0] else ""
            customer = safe_str(row[1]) if len(row) > 1 else ""
            sales = safe_str(row[4]) if len(row) > 4 else ""
            lot = safe_str(row[5]) if len(row) > 5 else ""
            dc = safe_str(row[6]) if len(row) > 6 else ""
            # '.'은 빈값 처리
            if lot == ".":
                lot = ""
            if dc == ".":
                dc = ""

            conn.execute(
                """INSERT INTO shipment_log
                   (ship_date, customer, part_number, quantity, sales_person, lot_number, datecode)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (ship_date, customer, pn, qty, sales, lot, dc),
            )
            inserted += 1

            # daily_inventory 출고 반영
            if ship_date:
                try:
                    d = datetime.strptime(ship_date, "%Y-%m-%d")
                    ym = d.strftime("%Y-%m")
                    day = d.day
                    conn.execute(
                        """INSERT INTO daily_inventory (part_number, year_month, day, outbound_qty)
                           VALUES (?, ?, ?, ?)
                           ON CONFLICT(part_number, year_month, day)
                           DO UPDATE SET outbound_qty = daily_inventory.outbound_qty + excluded.outbound_qty""",
                        (pn, ym, day, qty),
                    )
                except (ValueError, TypeError):
                    pass

    return {"total": total, "inserted": inserted, "errors": errors}
